"""Data export service (Excel / PDF / CSV)."""
import io
from datetime import datetime
from decimal import Decimal
from typing import List, Optional
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.ext.asyncio import AsyncSession

from app.config.constants import ExportFormat
from app.database.repositories.debt_repo import DebtRepository
from app.database.repositories.product_repo import ProductRepository
from app.database.repositories.transaction_repo import TransactionRepository
from app.schemas.export import ExportRequest, ExportResult
from app.services.base import BaseService


class ExportService(BaseService):
    """Service generating Excel / PDF / CSV reports."""

    def __init__(self, session: AsyncSession):
        super().__init__(session)
        self.tx_repo = TransactionRepository(session)
        self.debt_repo = DebtRepository(session)
        self.product_repo = ProductRepository(session)

    async def export_excel(self, request: ExportRequest) -> ExportResult:
        """Generate a styled multi-sheet Excel financial workbook."""
        wb = openpyxl.Workbook()
        ws_summary = wb.active
        ws_summary.title = "Xulosa"

        # Headers styling
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

        # 1. Summary Sheet
        ws_summary.append(["Biznes Moliya Hisoboti"])
        ws_summary.append(["Davr:", f"{request.start_date.strftime('%Y-%m-%d')} - {request.end_date.strftime('%Y-%m-%d')}"])
        ws_summary.append([])
        
        ws_summary.append(["Ko'rsatkich", "Qiymat"])
        ws_summary["A4"].font = header_font
        ws_summary["A4"].fill = header_fill
        ws_summary["B4"].font = header_font
        ws_summary["B4"].fill = header_fill

        # Transactions
        if request.include_transactions:
            ws_tx = wb.create_sheet(title="Tranzaksiyalar")
            ws_tx.append(["Sana", "Turi", "To'lov turi", "Summa (UZS)", "Tavsif"])
            for col in ["A1", "B1", "C1", "D1", "E1"]:
                ws_tx[col].font = header_font
                ws_tx[col].fill = header_fill

            transactions = await self.tx_repo.get_user_transactions(
                user_id=request.user_id,
                start_date=request.start_date,
                end_date=request.end_date,
                limit=5000,
            )
            for tx in transactions:
                ws_tx.append([
                    tx.transaction_date.strftime("%Y-%m-%d %H:%M"),
                    tx.type.value if hasattr(tx.type, "value") else str(tx.type),
                    tx.payment_method.value if hasattr(tx.payment_method, "value") else str(tx.payment_method),
                    float(tx.amount),
                    tx.description or "",
                ])

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        file_bytes = buffer.getvalue()

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"moliya_hisoboti_{timestamp}.xlsx"

        return ExportResult(
            file_bytes=file_bytes,
            filename=filename,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            size_bytes=len(file_bytes),
        )
